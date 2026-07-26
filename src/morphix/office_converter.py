"""Morphix office converter — XLSX/CSV pure Python + LibreOffice for Office formats."""

from __future__ import annotations

import csv
import logging
import shutil
import tempfile
from pathlib import Path
from typing import Any

from morphix.base import (
    BaseConverter,
    ConversionError,
    ConversionResult,
    ErrorCode,
    normalize_format,
)
from morphix.libreoffice import get_libreoffice

logger = logging.getLogger("verto.morphix.office")

# Conversions that require LibreOffice
_LO_PAIRS = {
    ("docx", "pdf"),
    ("pdf", "docx"),
    ("docx", "odt"),
    ("odt", "docx"),
    ("xlsx", "ods"),
    ("ods", "xlsx"),
    ("pptx", "pdf"),
    ("pdf", "pptx"),
    ("pptx", "odp"),
    ("odp", "pptx"),
    ("odt", "pdf"),
    ("ods", "pdf"),
    ("odp", "pdf"),
    ("doc", "pdf"),
    ("doc", "docx"),
    ("xls", "xlsx"),
    ("ppt", "pptx"),
}


class OfficeConverter(BaseConverter):
    """Office / spreadsheet conversions for Morphix."""

    name = "office"

    input_formats = {
        "docx",
        "odt",
        "pdf",
        "xlsx",
        "ods",
        "csv",
        "pptx",
        "odp",
        "doc",
        "xls",
        "ppt",
    }
    output_formats = {
        "docx",
        "odt",
        "pdf",
        "xlsx",
        "ods",
        "csv",
        "pptx",
        "odp",
    }

    def can_convert(self, source_format: str, target_format: str) -> bool:
        src = normalize_format(source_format)
        dst = normalize_format(target_format)
        if src == dst:
            return False
        # Pure Python paths
        if {src, dst} == {"xlsx", "csv"}:
            return True
        if (src, dst) in _LO_PAIRS:
            return True
        return False

    def list_targets(self, source_format: str) -> list[str]:
        src = normalize_format(source_format)
        targets: set[str] = set()
        if src == "xlsx":
            targets.update({"csv", "ods"})
        elif src == "csv":
            targets.add("xlsx")
        elif src == "ods":
            targets.update({"xlsx", "pdf"})
        elif src == "docx":
            targets.update({"pdf", "odt"})
        elif src == "odt":
            targets.update({"docx", "pdf"})
        elif src == "pptx":
            targets.update({"pdf", "odp"})
        elif src == "odp":
            targets.update({"pptx", "pdf"})
        elif src == "pdf":
            # PDF→DOCX/PPTX only via LO — also listed by pdf converter for images/txt
            targets.update({"docx"})
        elif src in {"doc", "xls", "ppt"}:
            mapping = {"doc": "docx", "xls": "xlsx", "ppt": "pptx"}
            targets.add(mapping[src])
            targets.add("pdf")
        return sorted(targets)

    def convert(
        self,
        source: Path,
        target_format: str,
        output_path: Path,
        options: dict[str, Any] | None = None,
    ) -> ConversionResult:
        options = options or {}
        source = Path(source)
        output_path = Path(output_path)
        src = normalize_format(source.suffix)
        dst = normalize_format(target_format)

        if not source.is_file():
            raise ConversionError(
                user_message="Source file not found.",
                technical_detail=str(source),
                code=ErrorCode.IO_ERROR,
            )

        output_path.parent.mkdir(parents=True, exist_ok=True)

        if src == "xlsx" and dst == "csv":
            return self._xlsx_to_csv(source, output_path, options)
        if src == "csv" and dst == "xlsx":
            return self._csv_to_xlsx(source, output_path, options)

        if (src, dst) in _LO_PAIRS or self.can_convert(src, dst):
            return self._via_libreoffice(source, dst, output_path)

        raise ConversionError(
            user_message=f"Unsupported office conversion: {src} → {dst}",
            code=ErrorCode.UNSUPPORTED,
        )

    def _xlsx_to_csv(
        self,
        source: Path,
        output_path: Path,
        options: dict[str, Any],
    ) -> ConversionResult:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ConversionError(
                user_message="openpyxl is required for XLSX conversion.",
                technical_detail=str(exc),
                code=ErrorCode.MISSING_DEPENDENCY,
            ) from exc

        sheet_index = int(options.get("sheet_index", 0))
        try:
            wb = load_workbook(source, read_only=True, data_only=True)
            try:
                if not wb.sheetnames:
                    raise ConversionError(
                        user_message="Workbook has no sheets.",
                        code=ErrorCode.EMPTY,
                    )
                if sheet_index < 0 or sheet_index >= len(wb.sheetnames):
                    sheet_index = 0
                ws = wb[wb.sheetnames[sheet_index]]
                with output_path.open("w", newline="", encoding="utf-8") as fh:
                    writer = csv.writer(fh)
                    for row in ws.iter_rows(values_only=True):
                        writer.writerow(["" if c is None else c for c in row])
            finally:
                wb.close()
        except ConversionError:
            raise
        except Exception as exc:  # noqa: BLE001
            raise ConversionError(
                user_message="Could not convert XLSX to CSV. The file may be corrupted.",
                technical_detail=str(exc),
                code=ErrorCode.CORRUPT,
            ) from exc

        return ConversionResult(
            output_path=output_path,
            source_path=source,
            target_format="csv",
            message="Converted XLSX to CSV",
        )

    def _csv_to_xlsx(
        self,
        source: Path,
        output_path: Path,
        options: dict[str, Any],
    ) -> ConversionResult:
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise ConversionError(
                user_message="openpyxl is required for CSV→XLSX conversion.",
                technical_detail=str(exc),
                code=ErrorCode.MISSING_DEPENDENCY,
            ) from exc

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            with source.open("r", newline="", encoding="utf-8", errors="replace") as fh:
                reader = csv.reader(fh)
                for row in reader:
                    ws.append(row)
            wb.save(output_path)
        except Exception as exc:  # noqa: BLE001
            raise ConversionError(
                user_message="Could not convert CSV to XLSX.",
                technical_detail=str(exc),
                code=ErrorCode.CORRUPT,
            ) from exc

        return ConversionResult(
            output_path=output_path,
            source_path=source,
            target_format="xlsx",
            message="Converted CSV to XLSX",
        )

    def _via_libreoffice(
        self,
        source: Path,
        target: str,
        output_path: Path,
    ) -> ConversionResult:
        lo = get_libreoffice()
        if not lo.available():
            raise ConversionError(
                user_message=(
                    "LibreOffice is required for this conversion but was not found. "
                    "Install LibreOffice (https://www.libreoffice.org) and ensure "
                    "'soffice' is available on your system PATH, then try again."
                ),
                code=ErrorCode.MISSING_DEPENDENCY,
            )

        with tempfile.TemporaryDirectory(prefix="verto-office-") as tmp:
            tmp_dir = Path(tmp)
            produced = lo.convert(source, target, output_dir=tmp_dir)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(produced, output_path)

        return ConversionResult(
            output_path=output_path,
            source_path=source,
            target_format=target,
            message=f"Converted via LibreOffice to {target.upper()}",
        )
