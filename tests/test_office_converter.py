"""Tests for Morphix OfficeConverter."""

from __future__ import annotations

from pathlib import Path

import pytest

from morphix.libreoffice import get_libreoffice
from morphix.office_converter import OfficeConverter


def test_xlsx_to_csv(fixtures_dir: Path, tmp_path: Path) -> None:
    conv = OfficeConverter()
    out = tmp_path / "out.csv"
    result = conv.convert(fixtures_dir / "sample.xlsx", "csv", out)
    text = result.output_path.read_text(encoding="utf-8")
    assert "alpha" in text
    assert "beta" in text


def test_csv_to_xlsx(fixtures_dir: Path, tmp_path: Path) -> None:
    conv = OfficeConverter()
    out = tmp_path / "out.xlsx"
    result = conv.convert(fixtures_dir / "sample.csv", "xlsx", out)
    assert result.output_path.is_file()
    from openpyxl import load_workbook

    wb = load_workbook(result.output_path)
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows[0][0] == "name"
    wb.close()


@pytest.mark.skipif(not get_libreoffice().available(), reason="LibreOffice not installed")
def test_csv_roundtrip_not_needed_lo() -> None:
    assert get_libreoffice().available()


@pytest.mark.skipif(not get_libreoffice().available(), reason="LibreOffice not installed")
def test_docx_to_pdf_via_lo(tmp_path: Path) -> None:
    """Create a minimal DOCX and convert to PDF with LibreOffice."""
    from docx import Document

    docx_path = tmp_path / "sample.docx"
    doc = Document()
    doc.add_paragraph("Hello from Morphix via LibreOffice")
    doc.save(docx_path)

    conv = OfficeConverter()
    out = tmp_path / "sample.pdf"
    result = conv.convert(docx_path, "pdf", out)
    assert result.output_path.is_file()
    assert result.output_path.stat().st_size > 0
